from dataclasses import dataclass
from typing import List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from ofx_parser import OFXParser, Transaction
from openpyxl import load_workbook
from datetime import datetime
import re
import shutil
import os

@dataclass
class OFX2AdaplRule:

    ofx_name: str
    name: str
    adapl_column: int
    vat_number: int

class Transaction2ADAPLWriter:

    def __init__(self, transactions, year, rules: List[OFX2AdaplRule], template_path="docs/ADAPLTemplate.xlsx", append: bool=False):
        self.transacs: List[Transaction] = transactions
        self.template_path: str = template_path
        self.year = year
        self.append = append
        self.path = self.copy_template()
        print(f"Open {self.path}")
        self.workbook = load_workbook(self.path)
        self.sheets_rows = self.find_first_empty_rows(13)
        self.sheets_positive_rows = self.find_first_empty_rows(1)
        self.rules: List[OFX2AdaplRule] = rules
        self.vats = [0.2, 0.55, 0]
        self.nb = 0

    def find_first_empty_rows(self, col):
        res = []
        for i in range(12):
            sheet = self.workbook.worksheets[i]
            row = 7
            while True:
                value = sheet.cell(row, col).value
                if value is None:
                    res.append(row)
                    break
                row+=1
        return res


    def copy_template(self):
        path = self.template_path.replace("Template", str(self.year - 2000))
        if not self.append:
            try:
                print(f"Creating {path}")
                shutil.copy(self.template_path, path)
            except PermissionError:
                input("Close Excel and press Enter")
                shutil.copy(self.template_path, path)
        return path

    def write(self):
        self.workbook.worksheets[0]["D1"] = self.year
        for t in self.transacs:
            self.write_transaction(t)
        if self.nb > 0:
            self.workbook.save(self.path)
            print(f"Saved {self.nb} transactions")
        else:
            print("No transaction to save")
            quit(0)

    def write_transaction(self, t:Transaction):
        dt = self.get_date(t.dt_user)
        sheet = self.workbook.worksheets[dt.month - 1]
        if not self.find_fit_id(t.fit_id, sheet):
            if t.trn_amt < 0:
                self.write_transaction_negative(t, dt, sheet)
            else:
                self.write_transaction_positive(t, dt, sheet)
            self.nb += 1

    def write_transaction_negative(self, t: Transaction, dt:datetime, sheet: Worksheet):
        row = self.sheets_rows[dt.month - 1]
        sheet.cell(row, 13).value = dt.strftime("%d/%m")
        rule = self.search_rule(t.name, -t.trn_amt)
        wt_col = 20
        vat_number = 0
        name = t.name
        if rule is not None:
            name = rule.name
            if rule.adapl_column != 0:
                wt_col = rule.adapl_column
            vat_number = rule.vat_number
        name = self.clean_name(name)
        sheet.cell(row, 14).value = name
        sheet.cell(row, 15).value = -t.trn_amt
        wt, vat = self.compute_vat(-t.trn_amt, vat_number)
        sheet.cell(row, 18).value = vat
        sheet.cell(row, wt_col).value = wt
        sheet.cell(row, 54).value = t.fit_id
        self.sheets_rows[dt.month - 1] += 1

    def write_transaction_positive(self, t: Transaction, dt:datetime, sheet: Worksheet):
        row = self.sheets_positive_rows[dt.month - 1]
        sheet.cell(row, 1).value = dt.strftime("%d/%m")
        name = self.clean_name(t.name)
        rule = self.search_rule(name, 0)
        if rule is not None:
            name = rule.name
        sheet.cell(row, 2).value = name
        sheet.cell(row, 3).value = t.trn_amt
        wt, vat = self.compute_vat(t.trn_amt, 0)
        sheet.cell(row, 6).value = wt
        if vat != 0:
            sheet.cell(row, 7).value = vat
        sheet.cell(row, 53).value = t.fit_id
        self.sheets_positive_rows[dt.month - 1] += 1

    def compute_vat(self, amount: float, vat_number):
        wt = round(amount / (1 + self.vats[vat_number]), 2)
        vat = amount - wt
        return wt, vat

    def clean_name(self, name):
        name = name.replace("PRLV ", "")
        name = name.replace("SEPA ", "")
        name = name.replace("VIR ", "")
        name = name.replace("VIREMENT", "")
        name = name.replace("CARTE ", "")
        name = name.replace("PAIEMENT PSC ", "")
        name = name.replace("PAIEMENT CB ", "")
        name = name.replace(" EUR", "")
        name = name.replace("VILLARD DE LA", "")
        name = name.replace("LANS EN VERCO", "")
        name = name.replace("PARIS", "")
        name = name.replace("MEYLAN", "")
        name = name.replace("GRENOBLE", "")
        name = name.replace("ST MARTIN D", "")
        name = name.replace("AUTRANS", "")
        name = re.sub("\d{4,99}", "", name)
        name = re.sub("\d{4,99}", "", name) # Not a bug
        if "*" in name:
            name = name[:name.index("*")]
        if name.endswith("/"):
            name = name[:-1]
        while "  " in name:
            name = name.replace("  ", " ").strip()
        return name


    def get_date(self, transaction_date:str) -> datetime:
        year = int(transaction_date[:4])
        month = int(transaction_date[4:6])
        day = int(transaction_date[6:])
        dt = datetime(year, month, day)
        return dt

    def search_rule(self, name, amount) -> Optional[OFX2AdaplRule]:
        for rule in self.rules:
            if rule.ofx_name in name:
                if rule.ofx_name == "HABITAT" and amount > 30:
                    rule.adapl_column = 49
                return rule
        if "PAIEMENT PSC" in name:
            return OFX2AdaplRule(name, name, 35, 1)
        return None

    def find_fit_id(self, id, sheet: Worksheet):
        for i in range(7, sheet.max_row + 1):
            if sheet.cell(i, 54).value == id:
                return True
            if sheet.cell(i, 53).value == id:
                return True
        return False


ofx2adapl_rules = [ #+12
    OFX2AdaplRule("INTERETS/FRAIS", "CREDIT MUTUEL", 41, 0),
    OFX2AdaplRule("DGFIP", "DGFIP ??", 24, 2),
    OFX2AdaplRule("SEGECO", "SEGECO", 41, 0),
    OFX2AdaplRule("IMPLID", "IMPLID", 41, 0),
    OFX2AdaplRule("MMA", "MMA", 33, 2),
    OFX2AdaplRule("AUTOMOBILE", "", 49, 2),
    OFX2AdaplRule("HABITAT", "CREDIT MUTUEL", 33, 2),
    OFX2AdaplRule("FRAIS", "CREDIT MUTUEL", 41, 2),
    OFX2AdaplRule("LOYER", "LOYER", 26, 2),
    OFX2AdaplRule("ELECTRICITE", "EDF ??", 49, 2),
    OFX2AdaplRule("FACTURE SGT", "CREDIT MUTUEL", 33, 0),
    OFX2AdaplRule("C.I.P.A.V", "CIPAV", 41, 0),
    OFX2AdaplRule("AREA", "AREA", 41, 0),
    OFX2AdaplRule("URSSAF", "URSSAF", 36, 2),
    OFX2AdaplRule("SFR", "SFR", 38, 0),
    OFX2AdaplRule("ORANGE", "ORANGE", 38, 0),
    OFX2AdaplRule("BOUYGUES", "BOUYGUES", 38, 0),
    OFX2AdaplRule("ZATOPEK", "ZATOPEK", 38, 1),
    OFX2AdaplRule("COMPTE", "", 49, 2),
    OFX2AdaplRule("LIVRET", "", 49, 2),
    OFX2AdaplRule("DAB", "", 49, 2),
    OFX2AdaplRule("C.A.M", "", 49, 2),
    OFX2AdaplRule("IKEA", "IKEA", 28, 0),
    OFX2AdaplRule("CDISCOUNT", "CDISCOUNT", 28, 0),
    OFX2AdaplRule("AWS", "AWS", 38, 0),
    OFX2AdaplRule("AMAZON", "AMAZON", 28, 0),
    OFX2AdaplRule("PAYPAL", "PAYPAL", 28, 0),

    OFX2AdaplRule("ATP", "ATP FORMATION", 0, 0),
    OFX2AdaplRule("BANQUE POPULAIRE", "BPAURA", 0, 0),
    OFX2AdaplRule("DICOM", "DICOM", 0, 0),
    OFX2AdaplRule("SKEMA", "SKEMA", 0, 0),
    OFX2AdaplRule("KERCIA", "KERCIA", 0, 0),
]




if __name__ == '__main__':

    print("OFX to ADAPL by Cyril Vincent")
    print("=============================")
    p = OFXParser()
    p.parse("docs/comptes2.ofx")
    print(p.transacs)
    w = Transaction2ADAPLWriter(p.transacs, 2022, ofx2adapl_rules, append=True)
    w.write()
    os.system(f"start EXCEL.EXE {w.path}")
